import datetime
import io
import json
import os
import time
import zipfile

import openpyxl as openpyxl
from django.db import transaction
from django.conf import settings
from django.contrib import messages
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.core.files.base import ContentFile
from django.http import HttpRequest, HttpResponse, JsonResponse, Http404
from django.shortcuts import render, HttpResponseRedirect, get_object_or_404
from django.urls import reverse
from geopy.geocoders import Nominatim
from selenium import webdriver

from .forms import OrderFileForm, OrderForm
from .map_funcs import get_map_screenshot, get_map, generate_docx, add_table
from .models import CurrentOrder, FulfilledOrder, FulfilledOrderImages, Region, Area as DB_Area, City, CurrentOrderFile, \
    PurposeBuilding, get_screenshot_path
from .permissions import IsOwnerOrReadOnly
from .rosreestr2 import GetArea
from .serializers import OrderSelializer
from .tasks import create_map_screenshot_task
from .validators import validate_number
from django.views.decorators.csrf import csrf_exempt


from pypdf import PdfReader
from rest_framework import generics
from io import StringIO

from rosreestr2coord import Area


class OrderDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = OrderSelializer
    queryset = FulfilledOrder.objects.all()
    permission_classes = (IsOwnerOrReadOnly,)


def region_autocomplete(request: HttpRequest) -> JsonResponse:
    if 'term' in request.GET:
        filtered_regions = Region.objects.filter(
            name__icontains=request.GET.get('term')
        )
        regions = []
        for region in filtered_regions:
            regions.append(region.name)

        return JsonResponse(regions, safe=False)


def area_autocomplete(request: HttpRequest) -> JsonResponse:
    region = Region.objects.get(
        name__icontains=request.GET.get('region')
    )
    areas = []
    for area in region.areas.all():
        areas.append(f'{region.name}, {area.name}')

    return JsonResponse(areas, safe=False)


def city_autocomplete(request: HttpRequest) -> JsonResponse:
    data = request.GET.get('region').split(', ')
    data.remove('')
    region, area = data
    areas = DB_Area.objects.get(name=area)
    citys = []
    for city in areas.citys.all():
        citys.append(f'{region}, {area}, {city.name}')

    return JsonResponse(citys, safe=False)


def ajax_validate_cadastral_number(request: HttpRequest) -> JsonResponse:
    cadastral_number = request.GET.get('cadastral_number', None)

    try:
        validate_number(cadastral_number)
        response = {
            'is_valid': True
        }
    except ValidationError:
        response = {
            'is_valid': False
        }

    return JsonResponse(response)


def ajax_get_coords(request: HttpRequest) -> JsonResponse:
    cadastral_number = request.GET.get('cadastral_number', None)

    try:
        validate_number(cadastral_number)
        area = Area(cadastral_number)
        coords = area.to_geojson_poly()

        response = {
            'is_valid': True,
            'coords': coords
        }

    except ValidationError:
        response = {
            'is_valid': False
        }

    return JsonResponse(response)


def ajax_get_coords_for_map_maker(request: HttpRequest) -> JsonResponse:
    if request.method != 'GET' or not request.is_ajax():
        return JsonResponse({'error': 'Invalid request'})

    cadastral_numbers = request.GET.getlist('cadastral_numbers[]')
    coords = {}
    for cadastral_number in cadastral_numbers:
        try:
            area = Area(cadastral_number)
            coords[cadastral_number] = area.to_geojson_poly()
        except:
            coords[cadastral_number] = None

    response = {
        'is_valid': True,
        'coords': coords
    }

    return JsonResponse(response)


# def ajax_get_coords_for_map_maker(request: HttpRequest) -> JsonResponse:
#     if request.method != 'GET' or not request.is_ajax():
#         return JsonResponse({'error': 'Invalid request'})
#
#     cadastral_numbers = request.GET.getlist('cadastral_numbers[]')
#     coords = []
#
#     for cadastral_number in cadastral_numbers:
#         try:
#             area = Area(cadastral_number)
#             coords.append(area.to_geojson_poly())
#         except:
#             response = {
#                 'is_valid': False
#             }
#
#     response = {
#         'is_valid': True,
#         'coords': coords
#     }
#
#     return JsonResponse(response)


def ajax_get_squares(request: HttpRequest) -> JsonResponse:
    if request.method != 'GET' or not request.is_ajax():
        return JsonResponse({'error': 'Invalid request'})

    unique_cadastral_numbers = request.GET.getlist('unique_cadastral_numbers[]')

    try:
        square_cadastral_area = [GetArea(i).attrs['area_value'] for i in unique_cadastral_numbers]
        square = sum(square_cadastral_area) / 10000

        response = {'is_valid': True, 'square': square}

    except ValidationError:
        response = {'is_valid': False}

    return JsonResponse(response)


def ajax_get_purpose_group(request: HttpRequest) -> JsonResponse:
    selected_value = request.GET.get('selected_value', None)

    try:
        purpose_building = PurposeBuilding.objects.get(purpose=selected_value)
        purpose_group = purpose_building.group.pk
        response = {
            'purpose_group': purpose_group
        }
    except ObjectDoesNotExist:
        response = {
            'error': 'Значение не найдено'
        }

    return JsonResponse(response)


def purpose_building_autocomplete(request):
    purpose_buildings = PurposeBuilding.objects.all().values_list(
        'purpose', flat=True
    )
    purpose_building_options = list(purpose_buildings)

    return JsonResponse(purpose_building_options, safe=False)


def view_index(request: HttpRequest) -> HttpResponse:
    objects = FulfilledOrder.objects.all()
    response = HttpResponseRedirect(reverse('expert:order'))

    request.session.modified = True
    try:
        request.session.pop('cadastral_numbers')
        request.session.pop('address')
    except KeyError:
        pass

    if 'cadastral_numbers' in request.POST:
        cadastral_numbers = request.POST.getlist('cadastral_numbers')
        request.session['cadastral_numbers'] = cadastral_numbers

        return response

    if 'address' in request.POST:
        address = request.POST.getlist('address')
        request.session['address'] = address

        return response

    if 'files' in request.FILES:
        files = request.FILES.getlist('files')
        cadastral_numbers = []
        for file in files:
            file_extension = os.path.splitext(file.name)[-1]
            if file_extension == '.pdf':
                reader = PdfReader(file)
                page = reader.pages[0]
                pdf_text = StringIO(page.extract_text())
                for text in pdf_text:
                    if 'Кадастровый номер' in text.strip():
                        cadastral = text.strip().split(' ')[-1]
                        cadastral_numbers.append(cadastral)
                request.session['cadastral_numbers'] = cadastral_numbers
            else:
                file = request.FILES.get('files').close()
                messages.error(request, 'Ошибка обработки файла')
                break

    return render(request, "geoexpert/index.html", context={'objects': objects})


@transaction.atomic
def view_order(request: HttpRequest) -> HttpResponse:
    # coordinates = []
    context = {}
    square_cadastral_area = []

    cadastral_numbers = request.session[
        'cadastral_numbers'
    ] if 'cadastral_numbers' in request.session else None

    address = request.session[
        'address'
    ] if 'address' in request.session else None

    if address:
        region, area, city = address[0].split(', ')

    if cadastral_numbers:
        cadastral_region = Region.objects.get(
            cadastral_region_number=cadastral_numbers[0].split(':')[0])
        cadastral_area = DB_Area.objects.filter(region=cadastral_region).get(
            cadastral_area_number=cadastral_numbers[0].split(':')[1])

        for number in cadastral_numbers:
            try:
                areas = GetArea(number)
                square_cadastral_area.append(areas.attrs['area_value'])
                # ?
                # coordinates += areas.get_coord()
            except KeyError:
                pass

    if request.method == 'POST':
        order_form = OrderForm(request.POST)
        order_files_form = OrderFileForm(request.POST, request.FILES)
        if order_form.is_valid() and order_files_form.is_valid():
            order = order_form.save()
            if cadastral_numbers:
                cadastral_numbers = request.POST.getlist('cadastral_numbers')
                new_cadastral_numbers = request.POST.getlist('new_cadastral_numbers')
                if new_cadastral_numbers:
                    cadastral_numbers += new_cadastral_numbers
                # order.coordinates = coordinates
                cadastral_numbers = list(filter(lambda x: not x.strip() == '', cadastral_numbers))
                order.cadastral_numbers = cadastral_numbers

                coordinates = get_coordinates(cadastral_numbers)
                order.coordinates = coordinates

                tmp_html = os.path.join(
                    settings.BASE_DIR, 'tmp', f'map-{order.id}.html')
                tmp_png = os.path.join(
                    settings.BASE_DIR, 'tmp', f'map-{order.id}.png')

                get_map_screenshot(order.cadastral_numbers).save(tmp_html)

                driver = webdriver.Chrome()
                driver.get(f'file://{tmp_html}')
                time.sleep(1)
                driver.save_screenshot(tmp_png)
                driver.quit()

                img_path = get_screenshot_path(order, 'map.png')
                with open(tmp_png, 'rb') as f:
                    order.map.save(img_path, ContentFile(f.read()), save=True)

                os.remove(tmp_html)
                os.remove(tmp_png)

            order.save()

            for file in request.FILES.getlist('file'):
                CurrentOrderFile.objects.create(order=order, file=file)
            messages.success(request, 'Ваша заявка отправлена')

            return HttpResponseRedirect(reverse('expert:index'))

    else:
        try:
            order_form = OrderForm(initial={
                'cadastral_numbers': cadastral_numbers if cadastral_numbers
                else None,

                'region': cadastral_region.id if cadastral_numbers
                else Region.objects.get(name=region).id,

                'area': cadastral_area.id if cadastral_numbers
                else DB_Area.objects.get(name=area).id,

                'city': None if cadastral_numbers
                else City.objects.get(name=city).id,

                'square_unit': CurrentOrder.SQUARE_UNIT[0][0],
            })
        except UnboundLocalError:
            order_form = OrderForm()

        order_files_form = OrderFileForm()

    context['squares'] = request.session[
        'square_from_mapmaker'
    ] if 'square_from_mapmaker' in request.session else square_cadastral_area
    context['order_form'] = order_form
    context['order_files_form'] = order_files_form
    context['purpose_building'] = PurposeBuilding.objects.all()
    context['cadastral_numbers'] = cadastral_numbers

    if 'square_from_mapmaker' in request.session:
        del request.session['square_from_mapmaker']

    return render(request, 'geoexpert/order.html', context=context)


def view_card(request: HttpRequest) -> HttpResponse:
    order = FulfilledOrderImages.objects.all().first()

    return render(request, 'geoexpert/card.html', context={'order': order})


def view_map_maker(request):
    return render(request, 'geoexpert/map_maker.html')


def view_order_pages(request: HttpRequest) -> HttpResponse:
    orders = CurrentOrder.objects.all().order_by(
        '-id').select_related('city', 'area', 'region', 'work_objective')
    context = {
        "orders": orders,
    }

    return render(request, 'geoexpert/order_pages.html', context=context)


def get_coordinates(cadastral_numbers: list) -> list:
    coordinates = []
    for number in cadastral_numbers:
        try:
            areas = GetArea(number)
            coordinates += areas.get_coord()
        except KeyError:
            pass
    return coordinates


def create_map_screenshot_view(request, order_id, cadastral_numbers):
    create_map_screenshot_task.delay(order_id, cadastral_numbers)
    return JsonResponse({'success': True})


def view_change_order_status(request: HttpRequest, order_id: int) -> HttpResponse:
    order = get_object_or_404(CurrentOrder.objects.select_related(
        'city', 'area', 'region', 'work_objective', 'user'),
        id=order_id)
    cadastral_numbers_lst = order.cadastral_numbers

    files = CurrentOrderFile.objects.filter(order=order)
    map_html = get_map(order.cadastral_numbers) if order.cadastral_numbers else False

    if request.method == 'POST':
        order_form = OrderForm(request.POST, instance=order)

        if order_form.is_valid():
            order.object_name = request.POST.get('object_name')
            old_cadastral = request.POST.getlist('cadastral_numbers')
            new_cadastral = request.POST.getlist('new_cadastral_numbers')

            if new_cadastral[0]:
                order.cadastral_numbers = old_cadastral + new_cadastral

                coordinates = get_coordinates(order.cadastral_numbers)
                order.coordinates = coordinates

                square_cadastral_area = []

                for i in order.cadastral_numbers:
                    areas = GetArea(i)
                    square_cadastral_area.append(areas.attrs['area_value'])
                if request.POST.get('square_unit') == "hectometer":
                    order.square = sum(square_cadastral_area) / 10000
                else:
                    order.square = sum(square_cadastral_area)

                order.save()
                create_map_screenshot_view(request, order_id, order.cadastral_numbers)

                cadastral_numbers_lst = order.cadastral_numbers
            else:
                order.cadastral_numbers = request.POST.getlist('cadastral_numbers')

                if cadastral_numbers_lst != order.cadastral_numbers:
                    coordinates = get_coordinates(order.cadastral_numbers)
                    order.coordinates = coordinates

                    square_cadastral_area = []

                    for i in order.cadastral_numbers:
                        areas = GetArea(i)
                        square_cadastral_area.append(areas.attrs['area_value'])
                    if request.POST.get('square_unit') == "hectometer":
                        order.square = sum(square_cadastral_area) / 10000
                    else:
                        order.square = sum(square_cadastral_area)

                    order.save()
                    create_map_screenshot_view(request, order_id, order.cadastral_numbers)

                    cadastral_numbers_lst = order.cadastral_numbers

            order_form.save()

            return JsonResponse({'success': True})
        else:
            errors_dict = {}
            for field, errors in order_form.errors.as_data().items():
                label = order_form.fields[field].label
                errors_dict[label] = [
                    {'label': label, 'message': error.message, 'code': error.code} for error in errors]
            return JsonResponse({'success': False, 'errors': errors_dict})
    else:
        order_form = OrderForm(instance=order)

    context = {
        'type_works': order.type_work.all(),
        'files': files,
        'order_form': order_form,
        'order': order,
        'map_html': map_html,
        'lengt_unit': order.get_length_unit_display()
    }

    return render(request, 'geoexpert/change_order_status.html', context=context)


def download_docx(request, document_name: str, document_path: str, document_cipher: str, placeholders: dict,
                  coordinates_dict: dict):
    document = generate_docx(document_path, placeholders)

    add_table(document, coordinates_dict)

    output = io.BytesIO()
    document.save(output)
    output.seek(0)

    document_name_upload = f'{document_cipher}-{document_name}'

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )
    response['Content-Disposition'] = f'attachment; filename="{document_name_upload}.docx"'

    return response


def download_igi_docx(request, pk: int):
    order = get_object_or_404(CurrentOrder, pk=pk)
    department = order.region.region_department.first()
    location = f"{order.region}, {order.area}, {order.city}, {order.street}, д.{order.house_number}"
    if order.building:
        location += f" {order.building}"

    date = datetime.datetime.now()
    document_cipher = f"{date.strftime('%Y%m%d')}-{order.pk:03d}"

    cadastral_numbers = order.cadastral_numbers
    coordinates = json.loads(order.coordinates)

    coordinates_dict = {}

    for i, coords in enumerate(coordinates):
        if i < len(cadastral_numbers):
            cadastral_num = cadastral_numbers[i]
            coordinates_dict[cadastral_num] = coords[0]

    document_name = 'IGI'
    document_path = os.path.join(settings.MEDIA_ROOT, f'{document_name}.docx')

    if department:
        placeholders = {
            '_шифр-иги': f'{document_cipher}-ИГИ',
            '_должность_руководителя_ведомства': department.director_position,
            '_название_ведомства': department.name,
            '_фио_руководителя_ведомства': f'{department.director_surname} {department.director_name} {department.director_patronymic}',
            '_тел_ведомства': str(department.phone_number),
            '_почта_ведомства': department.email,
            '_дата_текущая': date.strftime("%Y-%m-%d"),
            '_имя_руководителя_ведомства': department.director_name,
            '_название_объекта_полное': order.object_name,
            '_местоположение_объекта': location,
            '_кадастровый_номер': ', '.join(cadastral_numbers),
            '_шифр-тема': f'{document_cipher}-ИГИ'
        }
        if order.map:
            placeholders['_обзорная_схема'] = order.map.path
    else:
        placeholders = {}

    return download_docx(request, document_name,
                         document_path, document_cipher,
                         placeholders, coordinates_dict)


def download_igdi_docx(request, pk: int):
    order = get_object_or_404(CurrentOrder, pk=pk)
    department = order.region.region_department.first()
    location = f"{order.region}, {order.area}, {order.city}, {order.street}, д.{order.house_number}"
    if order.building:
        location += f" {order.building}"

    date = datetime.datetime.now()
    document_cipher = f"{date.strftime('%Y%m%d')}-{order.pk:03d}"

    cadastral_numbers = order.cadastral_numbers
    coordinates = json.loads(order.coordinates)

    coordinates_dict = {}

    for i, coords in enumerate(coordinates):
        if i < len(cadastral_numbers):
            cadastral_num = cadastral_numbers[i]
            coordinates_dict[cadastral_num] = coords[0]

    document_name = 'IGDI'
    document_path = os.path.join(settings.MEDIA_ROOT, f'{document_name}.docx')

    if department:
        placeholders = {
            '_шифр-игди': f'{document_cipher}-ИГДИ',
            '_должность_руководителя_ведомства': department.director_position,
            '_название_ведомства': department.name,
            '_фио_руководителя_ведомства': f'{department.director_surname} {department.director_name} {department.director_patronymic}',
            '_тел_ведомства': str(department.phone_number),
            '_почта_ведомства': department.email,
            '_дата_текущая': date.strftime("%Y-%m-%d"),
            '_имя_руководителя_ведомства': department.director_name,
            '_название_объекта_полное': order.object_name,
            '_местоположение_объекта': location,
            '_кадастровый_номер': ', '.join(cadastral_numbers),
            '_шифр-тема': f'{document_cipher}-ИГДИ'
        }
        if order.map:
            placeholders['_обзорная_схема'] = order.map.path
    else:
        placeholders = {}

    return download_docx(request, document_name,
                         document_path, document_cipher,
                         placeholders, coordinates_dict)


def download_all_docx(request, pk: int):
    igi_docx = download_igi_docx(request, pk)
    igdi_docx = download_igdi_docx(request, pk)

    document_cipher = f"{datetime.datetime.now().strftime('%Y%m%d')}-{pk:03d}"

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w") as archive:
        archive.writestr(f"{document_cipher}-igi.docx", igi_docx.content)
        archive.writestr(f"{document_cipher}-igdi.docx", igdi_docx.content)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = f"attachment; filename={document_cipher}.zip"
    return response


def download_map(request, pk: int):
    order = get_object_or_404(CurrentOrder, pk=pk)
    try:
        file_path = order.map.path
        with open(file_path, 'rb') as fh:
            response = HttpResponse(
                fh.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=' + \
                                              os.path.basename(file_path)
            return response
    except:
        raise Http404


def download_xlsx(request, pk: int):
    order = get_object_or_404(CurrentOrder, pk=pk)
    cadastral_numbers = order.cadastral_numbers

    coordinates_list = json.loads(order.coordinates)
    coordinates_dict = {}

    if len(cadastral_numbers) == 0:
        return HttpResponse('Нет кадастровых номеров для данного заказа')

    for i, coords in enumerate(coordinates_list):
        if i >= len(cadastral_numbers):
            break

        cadastral_num = cadastral_numbers[i]
        coordinates_dict[cadastral_num] = coords[0]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    current_row = 1

    for key, coords in coordinates_dict.items():
        sheet.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=3)
        sheet.cell(row=current_row, column=1,
                   value=f'Координаты углов участка {key}')
        current_row += 1

        sheet.cell(row=current_row, column=1, value='Номер точки')
        sheet.cell(row=current_row, column=2, value='Координата Х')
        sheet.cell(row=current_row, column=3, value='Координата У')
        current_row += 1

        for i, coord in enumerate(coords):
            sheet.cell(row=current_row, column=1, value=str(i + 1))
            sheet.cell(row=current_row, column=2, value=str(coord[0]))
            sheet.cell(row=current_row, column=3, value=str(coord[1]))
            current_row += 1

        current_row += 1

    document_cipher = f"{datetime.datetime.now().strftime('%Y%m%d')}-{order.pk:03d}"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=coord.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{document_cipher}-coord.xlsx"'
    workbook.save(response)

    return response


def ajax_get_address_by_coord(request: HttpRequest) -> JsonResponse:
    if request.method != 'GET' or not request.is_ajax():
        return JsonResponse({'error': 'Invalid request'})

    coords_center_lst = request.GET.get('coords')
    coords_list = json.loads(coords_center_lst)

    geolocator = Nominatim(user_agent="geoexpert")
    address = None

    for coords_center in coords_list:
        latitude = coords_center['lat']
        longitude = coords_center['lng']

        try:
            location = geolocator.reverse((latitude, longitude))
            address = location.raw['address']
        except Exception as e:
            print(f"Ошибка геокодирования ({latitude}, {longitude}): {str(e)}")
            continue

    if address:
        region = address.get('state', '')

        region_parts = region.split()
        capitalized_region_parts = [part for part in region_parts if part.istitle()]

        result = None

        if 'county' in address:
            district = address.get('county', '')
            district_parts = district.split()
            capitalized_district_parts = [part for part in district_parts if part.istitle()]

            for region_name in capitalized_region_parts:
                regions_db = Region.objects.filter(name__icontains=region_name)
                for region_obj in regions_db:
                    districts = region_obj.areas.filter(name__icontains=capitalized_district_parts[0])
                    for district_obj in districts:
                        result = {
                            'region': region_obj.name,
                            'district': district_obj.name
                        }

                        return JsonResponse({'address': address, 'result': result})
        else:
            for region_name in capitalized_region_parts:
                regions_db = Region.objects.filter(name__icontains=region_name)
                for region_obj in regions_db:
                    result = {
                        'region': region_obj.name
                    }
                    return JsonResponse({'address': address, 'result': result})
        return JsonResponse({'success': 'true', 'address': address})
    else:
        return JsonResponse({'error': 'No address found'})


def get_address_from_db(request: HttpRequest) -> JsonResponse:
    address = request.GET.get('address', None)

    region = address.get('state', '')
    district = address.get('county', '')
    locality = address.get('city', '') or address.get('town', '') or address.get('village', '')
    print(region, district, locality)

    region_db = Region.objects.get(name__icontains=region)
    area_db = DB_Area.objects.get(name__icontains=district)
    print(region_db, area_db)


def write_to_session_from_js(request):
    if request.method == 'POST':
        data = json.loads(request.body)

        for key, value in data.items():
            request.session[key] = value
            request.session.modified = True

        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

def save_translate_data(request):
    if request.method == "POST":
        new_datas = json.loads(request.body)
        with open("static/translate_data.json", "r", encoding="UTF-8") as file:
            existing_data = json.load(file)
            keys_intersection = new_datas.keys() & existing_data.keys()
            if not keys_intersection:
                existing_data.update(new_datas)
                with open("static/translate_data.json", "w", encoding="UTF-8") as file:
                    json.dump(existing_data, file, ensure_ascii=False, indent=2)
            else:
                return JsonResponse({"message": "Data already exists."})

        return JsonResponse({"message": "Data saved successfully."})
    else:
        return JsonResponse({"message": "Invalid request."}, status=400)
    